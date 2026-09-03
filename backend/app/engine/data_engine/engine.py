import csv
import json
import math
import re
from collections import Counter
from pathlib import Path
from typing import Any

from app.engine.core.base import DetectionEngine
from app.engine.core.context import DetectionContext
from app.engine.core.result import DetectionResult


REGEX_RULES = {
    "phone": re.compile(r"(?<!\d)1[3-9]\d{9}(?!\d)"),
    "id_card": re.compile(r"(?<!\d)[1-9]\d{5}(?:18|19|20)\d{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01])\d{3}[\dXx](?!\d)"),
    "bank_card": re.compile(r"(?<!\d)(?:62|4\d{3}|5[1-5]\d{2})[ -]?(?:\d[ -]?){12,17}(?!\d)"),
    "email": re.compile(r"\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b"),
    "api_key": re.compile(r"\b(?:AKIA|sk-[A-Za-z0-9]{20,}|ghp_[A-Za-z0-9]{20,}|AIza[0-9A-Za-z_-]{20,})\b"),
    "token": re.compile(r"\b(?:Bearer\s+)?[A-Za-z0-9_\-]{32,}\b"),
}


def shannon_entropy(text: str) -> float:
    if not text:
        return 0.0
    counter = Counter(text)
    length = len(text)
    return -sum((count / length) * math.log2(count / length) for count in counter.values())


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt", ".sql", ".log", ".json", ".xml", ".yaml", ".yml"}:
        return path.read_text(encoding="utf-8", errors="replace")
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(path, read_only=True, data_only=True)
            parts = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    parts.append(" ".join(str(value) for value in row if value is not None))
            return "\n".join(parts)
        except Exception:
            return ""
    if suffix == ".pdf":
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(str(path))
            return "\n".join((page.extract_text() or "") for page in reader.pages)
        except Exception:
            return ""
    return ""


def scan_text(text: str) -> dict[str, Any]:
    counts: dict[str, int] = {name: 0 for name in REGEX_RULES}
    samples: dict[str, list[str]] = {name: [] for name in REGEX_RULES}
    for name, pattern in REGEX_RULES.items():
        matches = pattern.findall(text)
        counts[name] = len(matches)
        samples[name] = [str(item) for item in matches[:10]]
    secret_candidates = [value for name in ("api_key", "token") for value in samples[name]]
    high_entropy = [value for value in secret_candidates if shannon_entropy(value) >= 3.5]
    return {
        "counts": counts,
        "samples": samples,
        "secret_count": counts["api_key"] + counts["token"],
        "pii_count": counts["phone"] + counts["id_card"] + counts["bank_card"] + counts["email"],
        "high_entropy_secrets": high_entropy[:20],
    }


def infer_columns(text: str, source: str) -> list[dict[str, Any]]:
    columns: list[str] = []
    if source.endswith(".csv"):
        try:
            reader = csv.reader(text.splitlines())
            columns = next(reader, [])
        except Exception:
            columns = []
    elif source.endswith(".sql"):
        columns = re.findall(r"^\s*([A-Za-z_][A-Za-z0-9_]*)\s+[A-Za-z0-9()]+", text, re.MULTILINE)
    elif source.endswith(".json"):
        try:
            data = json.loads(text)
            if isinstance(data, dict):
                columns = list(data.keys())
            elif isinstance(data, list) and data and isinstance(data[0], dict):
                columns = list(data[0].keys())
        except Exception:
            columns = []
    classified = []
    mappings = {
        "user_id": ["user_id", "userid", "uid"],
        "phone": ["phone", "mobile", "cellphone", "手机号", "联系电话"],
        "id_card": ["id_card", "idcard", "身份证", "identity_card"],
        "medical_record": ["medical_record", "病历", "patient_id", "诊断"],
    }
    for column in columns:
        categories = [category for category, keywords in mappings.items() if any(keyword in column.lower() for keyword in keywords)]
        classified.append({"name": column, "sensitivity": "High" if categories else "Unknown", "categories": categories})
    return classified


def presidio_scan(text: str) -> list[dict[str, str]]:
    from app.core.config import settings

    if not settings.presidio_enabled:
        return []
    try:
        from presidio_analyzer import AnalyzerEngine
        analyzer = AnalyzerEngine()
        results = analyzer.analyze(text=text, language="en")
        return [{"entity_type": item.entity_type, "score": round(float(item.score), 3), "start": item.start, "end": item.end} for item in results]
    except Exception:
        return []


def yara_scan(path: Path, rule_dir: Path) -> list[dict[str, Any]]:
    try:
        import yara
    except ImportError:
        return []
    rule_files = sorted(rule_dir.glob("*.yar")) if rule_dir.exists() else []
    if not rule_files:
        return []
    compiled = yara.compile(filepaths={file.stem: str(file) for file in rule_files})
    matches = compiled.match(str(path))
    return [{"rule": item.rule, "tags": item.tags, "meta": item.meta} for item in matches]


class DataEngine(DetectionEngine):
    name = "data_engine"
    version = "2.0.0"

    def analyze(self, context: DetectionContext) -> list[DetectionResult]:
        findings: list[DetectionResult] = []
        paths = list(context.files)
        if context.path:
            paths.append(context.path)
        for path in paths:
            if not path.exists():
                continue
            text = extract_text(path)
            scan = scan_text(text)
            presidio = presidio_scan(text) if text else []
            yara_matches = yara_scan(path, Path(__file__).resolve().parents[2] / "rules" / "data")
            evidence = {
                "file": str(path),
                "size": path.stat().st_size,
                "regex": scan,
                "presidio": presidio[:50],
                "yara": yara_matches[:50],
            }
            columns = infer_columns(text, path.name)
            if columns:
                evidence["columns"] = columns
                context.data.setdefault("data_assets", []).append({
                    "name": path.name,
                    "asset_type": "file",
                    "sensitivity": "High" if any(column["sensitivity"] == "High" for column in columns) else "Low",
                    "source": path.name,
                    "columns": columns,
                })
            if scan["pii_count"] > 0 or presidio:
                findings.append(DetectionResult(
                    engine=self.name,
                    rule_id="DATA_PII_001",
                    severity="High",
                    confidence=0.9 if scan["pii_count"] else 0.7,
                    evidence=evidence,
                    recommendation="对包含身份证、手机号、银行卡、邮箱等 PII 的文件实施加密、脱敏和访问控制。",
                ).normalize())
            if scan["secret_count"] > 0 or scan["high_entropy_secrets"]:
                findings.append(DetectionResult(
                    engine=self.name,
                    rule_id="DATA_SECRET_001",
                    severity="Critical",
                    confidence=0.85,
                    evidence=evidence,
                    recommendation="立即轮换泄露的密钥/Token，并排查代码、配置和备份文件。",
                ).normalize())
            if yara_matches:
                findings.append(DetectionResult(
                    engine=self.name,
                    rule_id="DATA_YARA_001",
                    severity="High",
                    confidence=0.9,
                    evidence=evidence,
                    recommendation="根据 YARA 规则检查文件来源、作者和是否包含恶意/敏感内容。",
                ).normalize())
        text = context.data.get("text", "")
        if text:
            scan = scan_text(text)
            if scan["pii_count"] > 0:
                findings.append(DetectionResult(
                    engine=self.name,
                    rule_id="DATA_PII_001",
                    severity="High",
                    confidence=0.9,
                    evidence={"text_sample": scan["samples"], "counts": scan["counts"]},
                    recommendation="对文本中的 PII 进行脱敏和最小化采集。",
                ).normalize())
            if scan["secret_count"] > 0:
                findings.append(DetectionResult(
                    engine=self.name,
                    rule_id="DATA_SECRET_001",
                    severity="Critical",
                    confidence=0.85,
                    evidence={"secret_samples": scan["samples"]["api_key"] + scan["samples"]["token"]},
                    recommendation="轮换泄露密钥，并从日志和配置中清除明文凭据。",
                ).normalize())
        return findings
